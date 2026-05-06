import streamlit as st
import pandas as pd
import io
import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

# =============================================================================
# CONFIGURAÇÕES DA PÁGINA
# =============================================================================
st.set_page_config(
    page_title="Consolida Workspace",
    layout="wide",
    initial_sidebar_state="expanded"
)

st.markdown("""
    <style>
    .main { background-color: #121212; }
    div.stButton > button:first-child { background-color: #5C2EE9; color: white; border-radius: 8px; width: 100%; }
    .stSelectbox, .stMultiSelect, .stNumberInput { color: white; }
    </style>
    """, unsafe_allow_html=True)

# =============================================================================
# FUNÇÕES DE UTILIDADE E EXTRAÇÃO
# =============================================================================
def formatar_data_excel_somente_data(val):
    try:
        if not val or val == "" or val == "0": return ""
        val_f = float(val)
        dt = datetime.datetime(1899, 12, 30) + datetime.timedelta(days=val_f)
        return dt.strftime('%d/%m/%Y')
    except (ValueError, TypeError):
        val_str = str(val).strip()
        if " " in val_str: val_str = val_str.split(" ")[0] 
        return val_str

def processar_arquivo_bruto(uploaded_file):
    try:
        if uploaded_file.name.endswith('.csv'):
            df_raw = pd.read_csv(uploaded_file, header=None, dtype=str).fillna("")
        else:
            df_raw = pd.read_excel(uploaded_file, header=None, dtype=str).fillna("")
        return df_raw.values.tolist()
    except Exception as e:
        st.error(f"Erro ao ler arquivo: {e}")
        return None

def extrair_dados_pre_conhecimento(linhas):
    dados_analiticos = []
    current_cte, current_emissao_cte, current_nf, current_emissao_nf = "", "", "", ""
    current_remetente, current_destinatario, current_peso, current_cub, current_valor = "", "", "", "", ""
    
    for i, row in enumerate(linhas):
        linha = [str(x).strip() for x in row]
        if not any(linha): continue
        if "Número" in linha and "CT-e" in linha:
            idx_cte = linha.index("CT-e")
            idx_emis_cte = linha.index("Emissão") if "Emissão" in linha else -1
            if i + 1 < len(linhas):
                sub = [str(x).strip() for x in linhas[i+1]]
                if len(sub) > idx_cte:
                    cands = [str(x).strip() for x in sub[idx_cte:idx_cte+5] if str(x).strip()]
                    if cands: current_cte = cands[0]
                if idx_emis_cte != -1 and len(sub) > idx_emis_cte:
                    cands_em = [str(x).strip() for x in sub[idx_emis_cte:idx_emis_cte+5] if str(x).strip()]
                    if cands_em: current_emissao_cte = formatar_data_excel_somente_data(cands_em[0])
                current_nf, current_emissao_nf, current_peso, current_cub, current_valor = "", "", "", "", ""

        for cell in linha:
            if "Remetente:" in cell:
                cands = [x for x in linha if x and "Remetente" not in x]
                if cands: current_remetente = cands[0].replace('\n', ' ')
            if "Destinatário:" in cell:
                cands = [x for x in linha if x and "Destinatário" not in x]
                if cands: current_destinatario = cands[0].replace('\n', ' ')
                
        if "Peso" in linha and "Cub." in linha and "Valor" in linha:
            idx_peso, idx_cub, idx_valor = linha.index("Peso"), linha.index("Cub."), linha.index("Valor")
            idx_emis_nf = linha.index("Emissão") if "Emissão" in linha else -1
            for j in range(i+1, min(i+5, len(linhas))):
                sub = [str(x).strip() for x in linhas[j]]
                if "NF" in sub:
                    idx_nf = sub.index("NF")
                    c_nf = [x for x in sub[idx_nf+1:] if x]
                    if c_nf: current_nf = c_nf[0]
                    if idx_emis_nf != -1 and len(sub) > idx_emis_nf:
                        c_em_nf = [x for x in sub[idx_emis_nf:idx_emis_nf+5] if x]
                        if c_em_nf: current_emissao_nf = formatar_data_excel_somente_data(c_em_nf[0])
                    try: current_peso = [x for x in sub[max(0, idx_peso-1):idx_peso+3] if x][0]
                    except: pass
                    try: current_cub = [x for x in sub[max(0, idx_cub-1):idx_cub+3] if x][0]
                    except: pass
                    try: current_valor = [x for x in sub[max(0, idx_valor-1):idx_valor+3] if x][0]
                    except: pass
                    break

        if "Frete calculado" in linha and "Frete realizado" in linha:
            idx_calc, idx_real = linha.index("Frete calculado"), linha.index("Frete realizado")
            j = i + 1
            while j < len(linhas):
                sub = [str(x).strip() for x in linhas[j]]
                if "Total do Frete" in sub or "Total de documentos" in sub or ("Número" in sub and "CT-e" in sub): break
                itens = [x for x in sub[:10] if x]
                if itens and len(sub) > max(idx_calc, idx_real):
                    nome = itens[0]
                    try: calc = float(sub[idx_calc]); real = float(sub[idx_real])
                    except: calc, real = 0.0, 0.0
                    diff = real - calc
                    status = "OK"
                    if diff > 0.01: status = "DIVERGÊNCIA (A MAIOR)"
                    elif diff < -0.01: status = "DIVERGÊNCIA (A MENOR)"
                    dados_analiticos.append({
                        "CT-e": current_cte, "Emissão CT-e": current_emissao_cte,
                        "NF": current_nf, "Emissão NF": current_emissao_nf,
                        "Remetente": current_remetente, "Destinatário": current_destinatario,
                        "Peso": current_peso, "Cub": current_cub, "Valor NF": current_valor,
                        "Componente": nome, "Calculado": calc, "Realizado": real, "Diferença": diff, "Status": status
                    })
                j += 1
    return dados_analiticos

def extrair_dados_embarque(linhas):
    dados_embarque = []
    cur_emb, cur_dt_criacao, cur_transp, cur_origem, cur_destino = "", "", "", "", ""
    for i, row in enumerate(linhas):
        linha = [str(x).strip() for x in row]
        if not any(linha): continue
        
        # CABEÇALHO DO EMBARQUE: Alterado para priorizar a coluna "Número" do espelho em vez de "Embarque" (ID)
        if "Embarque" in linha and "Transportadora" in linha:
            # Pega o Número se existir, senão faz fallback para Embarque
            idx_num_real = linha.index("Número") if "Número" in linha else linha.index("Embarque")
            idx_dt = linha.index("Dt. criação") if "Dt. criação" in linha else -1
            idx_transp = linha.index("Transportadora")
            
            if i + 1 < len(linhas):
                sub = [str(x).strip() for x in linhas[i+1]]
                if len(sub) > idx_num_real: 
                    cur_emb = sub[idx_num_real].replace(".0", "")
                if idx_dt != -1 and len(sub) > idx_dt:
                    cands_dt = [str(x).strip() for x in sub[idx_dt:idx_dt+5] if str(x).strip()]
                    if cands_dt: cur_dt_criacao = formatar_data_excel_somente_data(cands_dt[0])
                if len(sub) > idx_transp:
                    cands_tr = [str(x).strip() for x in sub[idx_transp:idx_transp+8] if str(x).strip()]
                    if cands_tr: cur_transp = cands_tr[0]
        
        if "Origem:" in str(row):
            cands = [x for x in linha if x and "Origem" not in x]
            if cands: cur_origem = cands[0].replace('\n', ' ')
        if "Destino:" in str(row):
            cands = [x for x in linha if x and "Destino" not in x]
            if cands: cur_destino = cands[0].replace('\n', ' ')

        if linha[0] == "Nome" and "Frete calculado" in linha and "Frete realizado" in linha:
            idx_calc, idx_real = linha.index("Frete calculado"), linha.index("Frete realizado")
            j = i + 1
            while j < len(linhas):
                sub = [str(x).strip() for x in linhas[j]]
                if not any(sub) or "Total" in sub[0] or "Pré-conhecimentos" in sub or "Embarque" in sub: break
                nome = sub[0]
                if nome and len(sub) > max(idx_calc, idx_real):
                    try: calc, real = float(sub[idx_calc]), float(sub[idx_real])
                    except: calc, real = 0.0, 0.0
                    dados_embarque.append({
                        "Embarque ID": cur_emb, "Data Criação": cur_dt_criacao,
                        "Transportadora": cur_transp, "Origem": cur_origem, "Destino": cur_destino,
                        "Componente": nome, "Calculado": calc, "Realizado": real, "Diferença": real - calc
                    })
                j += 1
    return dados_embarque

# =============================================================================
# LÓGICA DE STATUS E FORMATAÇÃO DE OBS
# =============================================================================
def definir_status(diff, tolerancia):
    if abs(diff) <= tolerancia: return "OK"
    return "DIVERGÊNCIA (A MAIOR)" if diff > tolerancia else "DIVERGÊNCIA (A MENOR)"

def formatar_linha_observacao(row):
    status = row['Status']
    sufixo = "Divergencia a Maior" if "A MAIOR" in status else "Divergencia a Menor"
    return f"{row['Componente']} - {sufixo}"

# =============================================================================
# GERADORES DE EXCEL
# =============================================================================
def gerar_excel_colorido(df_local):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_local.to_excel(writer, index=False, sheet_name='Auditoria')
        ws = writer.sheets['Auditoria']
        
        fill_cab = PatternFill(start_color="5C2EE9", end_color="5C2EE9", fill_type="solid")
        font_cab = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = fill_cab; cell.font = font_cab

        fill_verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        fill_vermelho = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        fill_amarelo = PatternFill(start_color="F4D03F", end_color="F4D03F", fill_type="solid")

        cols = {col: i + 1 for i, col in enumerate(df_local.columns)}
        start_col = cols.get('Componente', 1)
        end_col = cols.get('Status', len(df_local.columns))
        
        for row_idx, row_data in enumerate(df_local.itertuples(), start=2):
            status = str(row_data.Status).upper()
            target_fill = fill_verde if "OK" in status else (fill_vermelho if "A MAIOR" in status else fill_amarelo)
            for col_idx in range(start_col, end_col + 1):
                ws.cell(row=row_idx, column=col_idx).fill = target_fill

        ws.auto_filter.ref = ws.dimensions
        for col in ws.columns:
            max_len = max([len(str(cell.value)) for cell in col])
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)
    return output.getvalue()

def gerar_excel_unificado_embarque(df_final):
    output = io.BytesIO()
    
    df_obs_input = df_final[df_final['Status'] != "OK"].copy()
    if not df_obs_input.empty:
        df_obs_input['Linha_Formatada'] = df_obs_input.apply(formatar_linha_observacao, axis=1)
        df_resumo = df_obs_input.groupby('Embarque ID')['Linha_Formatada'].apply(lambda x: " | ".join(x.unique())).reset_index()
        df_resumo.columns = ['Embarque', 'Observação']
    else:
        df_resumo = pd.DataFrame(columns=['Embarque', 'Observação'])

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_final.to_excel(writer, index=False, sheet_name='Analítico Detalhado')
        df_resumo.to_excel(writer, index=False, sheet_name='Resumo Observações')
        
        ws1, ws2 = writer.sheets['Analítico Detalhado'], writer.sheets['Resumo Observações']
        fill_cab = PatternFill(start_color="5C2EE9", end_color="5C2EE9", fill_type="solid")
        font_cab = Font(color="FFFFFF", bold=True)
        fill_vermelho = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        fill_amarelo = PatternFill(start_color="F4D03F", end_color="F4D03F", fill_type="solid")
        fill_verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

        for cell in ws1[1]: cell.fill, cell.font = fill_cab, font_cab
        for row_idx, row_data in enumerate(df_final.itertuples(), start=2):
            status = str(row_data.Status)
            color = fill_verde if status == "OK" else (fill_vermelho if "A MAIOR" in status else fill_amarelo)
            for col_idx in range(1, len(df_final.columns) + 1):
                ws1.cell(row=row_idx, column=col_idx).fill = color

        for cell in ws2[1]: cell.fill, cell.font = fill_cab, font_cab
        ws2.column_dimensions['A'].width, ws2.column_dimensions['B'].width = 20, 100

    return output.getvalue()

# =============================================================================
# INTERFACE STREAMLIT
# =============================================================================
st.sidebar.markdown("# 📊 CONSOLIDA\n### WORKSPACE")
st.sidebar.divider()
modulo = st.sidebar.radio("Navegação", ["Auditoria de Frete"])

if modulo == "Auditoria de Frete":
    st.title("Módulo de Extração Analítica")
    
    # -------------------------------------------------------------------------
    # CRIAÇÃO DAS 3 ABAS
    # -------------------------------------------------------------------------
    tab_cte, tab_emb, tab_cruzamento = st.tabs(["📦 Pré-Conhecimentos", "🚢 Embarques Globais", "🔗 Cruzamento Auditoria"])

    # --- ABA 1: CT-E ---
    with tab_cte:
        arquivo_cte = st.file_uploader("Upload CT-e", type=['xlsx', 'csv'], key="u_cte")
        if arquivo_cte and st.button("🚀 Analisar Arquivo CT-e"):
            linhas = processar_arquivo_bruto(arquivo_cte)
            if linhas:
                dados_cte = extrair_dados_pre_conhecimento(linhas)
                if dados_cte:
                    df_cte = pd.DataFrame(dados_cte)
                    st.dataframe(df_cte, use_container_width=True)
                    excel_cte = gerar_excel_colorido(df_cte)
                    st.download_button("⬇️ Baixar Excel Analítico (CT-e)", data=excel_cte, file_name="Auditoria_CTE.xlsx")

    # --- ABA 2: EMBARQUES ---
    with tab_emb:
        arquivo_emb = st.file_uploader("Upload Embarques", type=['xlsx', 'csv'], key="u_emb")
        if arquivo_emb:
            if st.button("🚀 Analisar Arquivo de Embarques"):
                linhas = processar_arquivo_bruto(arquivo_emb)
                if linhas: st.session_state['dados_emb_brutos'] = extrair_dados_embarque(linhas)

            if 'dados_emb_brutos' in st.session_state:
                df_base = pd.DataFrame(st.session_state['dados_emb_brutos'])
                
                st.divider()
                st.subheader("⚙️ Parâmetros de Refino")
                c1, c2, c3 = st.columns([2, 2, 1])
                
                with c1:
                    sel_comp = st.multiselect("Flegar Componentes:", options=df_base['Componente'].unique().tolist(), default=df_base['Componente'].unique().tolist())
                with c2:
                    sel_div = st.selectbox("Tipo de Filtro Exportação/Tela:", ["Todas", "Divergências", "A Maior", "A Menor"])
                with c3:
                    tolerancia = st.number_input("Tolerância (R$):", min_value=0.0, value=0.01, step=0.01)

                df_final = df_base[df_base['Componente'].isin(sel_comp)].copy()
                df_final['Status'] = df_final['Diferença'].apply(lambda x: definir_status(x, tolerancia))

                if sel_div == "Divergências": df_final = df_final[df_final['Status'] != "OK"]
                elif sel_div == "A Maior": df_final = df_final[df_final['Status'] == "DIVERGÊNCIA (A MAIOR)"]
                elif sel_div == "A Menor": df_final = df_final[df_final['Status'] == "DIVERGÊNCIA (A MENOR)"]

                st.write(f"**Registros Aplicados:** {len(df_final)}")
                st.dataframe(df_final, use_container_width=True)

                if not df_final.empty:
                    st.divider()
                    excel_data = gerar_excel_unificado_embarque(df_final)
                    st.download_button(
                        label="⬇️ Baixar Relatório Unificado (Analítico + Observações)",
                        data=excel_data,
                        file_name=f"Relatorio_Embarque_{datetime.date.today()}.xlsx"
                    )

    # --- ABA 3: CRUZAMENTO CORRIGIDO ---
    with tab_cruzamento:
        st.info("Cruze o relatório gerado na aba anterior com o Relatório Completo do sistema para obter a visão unificada.")
        col_cruz1, col_cruz2 = st.columns(2)
        
        with col_cruz1:
            arq_divergencias = st.file_uploader("1️⃣ Arquivo Gerado (Aba Resumo Observações)", type=['xlsx'], key="u_div")
        with col_cruz2:
            arq_relatorio = st.file_uploader("2️⃣ Relatório do Sistema (Embarque na Coluna T)", type=['xlsx', 'csv'], key="u_rel")
            
        if arq_divergencias and arq_relatorio:
            if st.button("🔗 Processar Cruzamento (PROCV)"):
                try:
                    with st.spinner("Lendo arquivos e executando cruzamento..."):
                        # Leitura da aba de Observações
                        try:
                            df_resumo = pd.read_excel(arq_divergencias, sheet_name='Resumo Observações', dtype=str)
                        except:
                            df_resumo = pd.read_excel(arq_divergencias, dtype=str)
                            
                        if 'Embarque' not in df_resumo.columns or 'Observação' not in df_resumo.columns:
                            st.error("O primeiro arquivo não possui as colunas esperadas. Use o arquivo unificado baixado na Aba 2.")
                        else:
                            # Leitura do sistema
                            if arq_relatorio.name.endswith('.csv'):
                                df_sistema = pd.read_csv(arq_relatorio, dtype=str)
                            else:
                                df_sistema = pd.read_excel(arq_relatorio, dtype=str)

                            if len(df_sistema.columns) >= 20:
                                nome_coluna_t = df_sistema.columns[19] # Índice 19 corresponde à Coluna T
                                
                                # FUNÇÃO AGRESSIVA DE LIMPEZA PARA GARANTIR MATCH DO PROCV
                                def padronizar_chave(valor):
                                    if pd.isna(valor): return ""
                                    v = str(valor).strip().upper()
                                    if v.endswith('.0'): v = v[:-2]
                                    if v.isdigit(): return str(int(v)) # Remove Zeros à esquerda (Ex: 0090075 -> 90075)
                                    return v
                                    
                                # Cria colunas temporárias invisíveis apenas para fazer o vínculo
                                df_sistema['_chave_procv'] = df_sistema.iloc[:, 19].apply(padronizar_chave)
                                df_resumo['_chave_procv'] = df_resumo['Embarque'].apply(padronizar_chave)
                                
                                # Renomeia a observação para não causar conflito com colunas já existentes no sistema
                                df_resumo.rename(columns={'Observação': 'Observação Auditoria'}, inplace=True)

                                # Executa o Merge (PROCV)
                                df_cruzado = pd.merge(df_sistema, df_resumo[['_chave_procv', 'Observação Auditoria']], 
                                                      on='_chave_procv', how='left')
                                
                                # DELETA APENAS A COLUNA TEMPORÁRIA. Nenhuma coluna original é afetada.
                                df_cruzado.drop(columns=['_chave_procv'], inplace=True)
                                
                                # Preenche dados sem divergência com "-"
                                df_cruzado['Observação Auditoria'] = df_cruzado['Observação Auditoria'].fillna("-")
                                
                                st.success(f"Cruzamento concluído com sucesso usando a coluna '{nome_coluna_t}'!")
                                st.dataframe(df_cruzado.head(50), use_container_width=True)

                                # Exportação
                                output_cruz = io.BytesIO()
                                with pd.ExcelWriter(output_cruz, engine='openpyxl') as writer:
                                    df_cruzado.to_excel(writer, index=False, sheet_name='Base Final')
                                    ws_cruz = writer.sheets['Base Final']
                                    
                                    fill_cab = PatternFill(start_color="5C2EE9", end_color="5C2EE9", fill_type="solid")
                                    font_cab = Font(color="FFFFFF", bold=True)
                                    for cell in ws_cruz[1]:
                                        cell.fill = fill_cab; cell.font = font_cab

                                st.download_button(
                                    label="⬇️ Baixar Base Final Cruzada",
                                    data=output_cruz.getvalue(),
                                    file_name=f"Resultado_Final_TMS_{datetime.date.today()}.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                )
                            else:
                                st.error(f"O Relatório do sistema tem apenas {len(df_sistema.columns)} colunas. A Coluna T é a 20ª coluna, o arquivo está fora do padrão.")
                except Exception as e:
                    st.error(f"Erro no processamento: {e}")
